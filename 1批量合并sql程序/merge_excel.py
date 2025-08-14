# merge_excel.py

import pandas as pd
import os
import warnings
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_00


def merge_excel_files(directory, file_order, output_file):
    """
    从指定目录读取Excel文件，并合并数据。
    
    参数:
    directory (str): 包含Excel文件的目录路径。
    file_order (list): 需要按顺序读取的文件名列表（不包含文件扩展名）。
    output_file (str): 输出合并后的Excel文件的路径。
    """
    # 构建完整的文件路径列表
    excel_files = [os.path.join(directory, f'{rule}.xlsx') for rule in file_order]

    # 初始化一个空的DataFrame列表
    dataframes = []
    province_dfs = []

    # 按顺序读取文件，并将数据添加到dataframes列表中
    for idx, file in enumerate(excel_files):
        if os.path.exists(file):
            # 读取省份列
            province_df = pd.read_excel(file, usecols=['省份'])
            province_dfs.append(province_df)

            # 读取其他列，并存储到dataframes中
            df = pd.read_excel(file, usecols=['稽核成功数量', '稽核总量', '汇总成功率'])
            
            # 为每列加上前缀，确保列名唯一
            df.columns = [f"{col}_{idx}" for col in df.columns]
            dataframes.append(df)
        else:
            print(f"文件 {file} 不存在，将跳过。")

    # 合并所有省份信息
    all_provinces = pd.concat(province_dfs).drop_duplicates().reset_index(drop=True)

    # 确保至少有一个文件被正确读取
    if not dataframes:
        print("没有可合并的数据")
    else:
        # 创建一个空的DataFrame，包含所有省份作为索引
        merged_df = pd.DataFrame(index=all_provinces['省份'])

        # 按照文件顺序填充数据
        for idx, df in enumerate(dataframes):
            # 设置省份作为索引
            df.set_index(province_dfs[idx]['省份'], inplace=True)
            
            # 确保索引唯一
            if not df.index.is_unique:
                df = df.reset_index().drop_duplicates(subset='省份').set_index('省份')
            
            # 使用省份索引填充数据
            merged_df = pd.concat([merged_df, df], axis=1)

        # 重置索引，使省份成为一列
        merged_df.reset_index(inplace=True)
        merged_df.rename(columns={'index': '省份'}, inplace=True)

        # 添加新的汇总列，并跳过空值
        merged_df['稽核成功数量之和'] = merged_df.filter(regex='^稽核成功数量').sum(axis=1, skipna=True)
        merged_df['稽核总量之和'] = merged_df.filter(regex='^稽核总量').sum(axis=1, skipna=True)
        merged_df['总汇总成功率'] = (merged_df['稽核成功数量之和'] / merged_df['稽核总量之和']).fillna(0) * 100

        # 格式化“总汇总成功率”为百分比，并处理 NaN 值
        merged_df['总汇总成功率'] = merged_df['总汇总成功率'].apply(lambda x: '{:.2f}%'.format(x) if pd.notna(x) else '')

        # 对“总汇总成功率”列进行排序
        merged_df.sort_values(by='总汇总成功率', ascending=True, inplace=True, na_position='last')

        # 提取“全国”的行
        national_row = merged_df[merged_df['省份'] == '全国']
        # 移除“全国”的行
        merged_df = merged_df[merged_df['省份'] != '全国']

        # 将“全国”的行添加到末尾
        merged_df = pd.concat([merged_df, national_row])

        # 忽略警告信息
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")

            # 创建一个新的工作簿
            wb = Workbook()
            ws = wb.active

            # 设置列标题
            for col_num, header in enumerate(merged_df.columns, 1):
                ws.cell(row=1, column=col_num, value=header).alignment = Alignment(horizontal="center")

            # 将DataFrame写入工作表
            for r_idx, row in enumerate(merged_df.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        if isinstance(value, str) and '%' in value:
                            # 设置百分比格式
                            cell.value = float(value[:-1])  # 去掉百分号转换成浮点数
                            cell.number_format = FORMAT_PERCENTAGE_00
                        else:
                            # 设置数值格式
                            cell.number_format = FORMAT_NUMBER_00
                        
                    # 对齐居中
                    cell.alignment = Alignment(horizontal="center")

            # 保存工作簿
            wb.save(output_file)

        print(f"合并后的Excel文件已保存到 {output_file}")