# 查找相对上月，下降的省份

import pandas as pd
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

def process_files(directory_a, directory_b, output_directory):
    # 获取目录 A 和 B 中的所有 Excel 文件
    files_a = [f for f in os.listdir(directory_a) if f.endswith('.xlsx')]
    files_b = [f for f in os.listdir(directory_b) if f.endswith('.xlsx')]

    # 截取文件名称中 "_" 之前的文案
    names_a = {os.path.splitext(f)[0].split('_')[0]: f for f in files_a}
    names_b = {os.path.splitext(f)[0].split('_')[0]: f for f in files_b}

    # 找到两个目录下相同的文案
    common_names = set(names_a.keys()) & set(names_b.keys())

    # 创建输出文件夹（如果不存在）
    output_subdirectory = os.path.join(output_directory, "省份成功率变化")
    if not os.path.exists(output_subdirectory):
        os.makedirs(output_subdirectory)

    # 遍历相同文案
    for name in common_names:
        # 构建文件路径
        file_a = os.path.join(directory_a, names_a[name])
        file_b = os.path.join(directory_b, names_b[name])

        # 读取省份和总汇总成功率列
        df_a = pd.read_excel(file_a, usecols=['省份', '总汇总成功率'])
        df_b = pd.read_excel(file_b, usecols=['省份', '总汇总成功率'])

        # 将百分比字符串转换为浮点数
        df_a['总汇总成功率'] = df_a['总汇总成功率'].apply(lambda x: float(x.strip('%')) / 100 if isinstance(x, str) and '%' in x else x)
        df_b['总汇总成功率'] = df_b['总汇总成功率'].apply(lambda x: float(x.strip('%')) / 100 if isinstance(x, str) and '%' in x else x)

        # 重命名列以区分来源
        df_b.rename(columns={'总汇总成功率': '上个月总汇总成功率'}, inplace=True)

        # 合并数据
        merged_df = pd.merge(df_a, df_b, on='省份', how='left')

        # 计算成功率变化数值
        merged_df['成功率变化数值'] = merged_df['总汇总成功率'] - merged_df['上个月总汇总成功率']

        # 生成新的 Excel 文件
        current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(output_subdirectory, f"{name}_省份成功率变化查值_{current_time}.xlsx")

        # 使用 pandas 保存 Excel 文件
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            merged_df.to_excel(writer, index=False, sheet_name='Sheet1')

        # 使用 openpyxl 进行额外的格式化
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

        # 设置列标题
        for col_num, header in enumerate(merged_df.columns, 1):
            ws.cell(row=1, column=col_num, value=header).alignment = Alignment(horizontal="center", vertical="center")

        # 将 DataFrame 写入工作表
        for r_idx, row in enumerate(merged_df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(value, float) and value >= 0 and value <= 1:
                    cell.number_format = FORMAT_PERCENTAGE_00
        
        # 保存工作簿
        wb.save(output_file)

        print(f"文件已保存到 {output_file}")
# D:\智网创新\2月报\月报ppt\12月
directory_a = r"D:\智网创新\2月报\月报ppt\1月\1月sql合并后的数据"
directory_b = r"D:\智网创新\2月报\月报ppt\1月\12月sql合并后的数据"
output_directory = directory_a

process_files(directory_a, directory_b, output_directory)