# 查找相对上月，稽核成功率的变化
import pandas as pd
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

def process_files(directory_a, directory_b, output_directory):
    # 获取目录 A 和 B 中的所有 Excel 文件
    files_a = [f for f in os.listdir(directory_a) if f.endswith('.xlsx')]
    files_b = [f for f in os.listdir(directory_b) if f.endswith('.xlsx')]

    # 截取文件名称中 "_" 之前的文案
    names_a = {os.path.splitext(f)[0].split('_')[0]: f for f in files_a}
    names_b = {os.path.splitext(f)[0].split('_')[0]: f for f in files_b}

    # 找到两个目录下相同的文案
    common_names = set(names_a.keys()) & set(names_b.keys())

    # 创建输出文件夹（如果不存在）
    output_subdirectory = os.path.join(output_directory, "省份稽核规则成功率变化")
    if not os.path.exists(output_subdirectory):
        os.makedirs(output_subdirectory)

    # 遍历相同文案
    for name in common_names:
        # 构建文件路径
        file_a = os.path.join(directory_a, names_a[name])
        file_b = os.path.join(directory_b, names_b[name])

        # 读取所有列信息
        df_a = pd.read_excel(file_a)
        df_b = pd.read_excel(file_b)

        # 只保留需要的列
        cols_a = ['省份'] + [col for col in df_a.columns if '成功率' in col]
        cols_b = ['省份'] + [col for col in df_b.columns if '成功率' in col]

        # 确保两个文件中的列数一致
        if len(cols_a) != len(cols_b):
            print(f"警告：文件 {name} 列数不一致，跳过该文件")
            continue

        # 提取需要的列
        df_a = df_a[cols_a]
        df_b = df_b[cols_b]

        # 将百分比字符串转换为浮点数
        for col in df_a.columns:
            if '成功率' in col:
                df_a[col] = df_a[col].apply(lambda x: float(x.strip('%')) / 100 if isinstance(x, str) and '%' in x else x)
        for col in df_b.columns:
            if '成功率' in col:
                df_b[col] = df_b[col].apply(lambda x: float(x.strip('%')) / 100 if isinstance(x, str) and '%' in x else x)

        # 重新索引以确保列顺序一致
        df_b.columns = df_a.columns

        # 根据“省份”列合并数据
        merged_df = pd.merge(df_a, df_b, on='省份', how='left', suffixes=('', '_b'))

        # 计算所有“成功率”列的差值
        success_rate_cols = [col for col in df_a.columns if '成功率' in col]

        for col in success_rate_cols:
            new_col_name = f'{col}差值'
            merged_df[new_col_name] = merged_df[col] - merged_df[col + '_b']

        # 生成新的 DataFrame，只包含省份列和差值列
        result_df = merged_df[['省份'] + [f'{col}差值' for col in success_rate_cols]]

        # 生成新的 Excel 文件
        current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(output_subdirectory, f"{name}_成功率差值_{current_time}.xlsx")

        # 使用 pandas 保存 Excel 文件
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Sheet1')

        # 使用 openpyxl 进行额外的格式化
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

        # 设置列标题
        for col_num, header in enumerate(result_df.columns, 1):
            ws.cell(row=1, column=col_num, value=header).alignment = Alignment(horizontal="center", vertical="center")

        # 将 DataFrame 写入工作表
        for r_idx, row in enumerate(result_df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(value, float) and value >= 0 and value <= 1:
                    cell.number_format = FORMAT_PERCENTAGE_00

        # 保存工作簿
        wb.save(output_file)

        print(f"文件已保存到 {output_file}")

# 指定目录D:\智网创新\2月报\月报ppt\12月
directory_a = r"D:\智网创新\2月报\月报ppt\1月\1月sql合并后的数据"
directory_b = r"D:\智网创新\2月报\月报ppt\1月\12月sql合并后的数据"
output_directory = directory_a

# 调用函数
process_files(directory_a, directory_b, output_directory)